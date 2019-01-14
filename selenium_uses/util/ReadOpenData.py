
from openpyxl import load_workbook
import os
class ReadOpenData:
    """
    读取配置类
    """

    path =os.path.dirname(os.path.abspath(os.path.dirname(__file__)))
    def __init__(self):
        "Constructor of the object"
        self.workbook = load_workbook(self.path+os.sep+"test_data"+os.sep+"data.xlsx")

    def get_sheet(self):
        "读取sheet部分"
        return self.workbook.sheetnames[0]

    def get_cell_dire(self,row,col):
        """
        通过坐标直接读取单元格
        :param row: int
        :param col: int
        :return: 对应row & col 的 cell
        """
        row+=1
        col+=1
        return self.workbook.active.cell(row, col).value

    def get_rows(self,row):
        """
        读取某一行数据
        :param row: int
        :return:包含这一行数据的list  class:openpyxl.cell
        """
        values = []
        cells = list(self.workbook.active.iter_rows())[row]
        for cell in cells:
            values.append(cell.value)
        return values


    def get_cell_by_row(self,row,col):
        """
        按行读取单元格
        :param row: int
        :param col: int
        :return: str, 对应row & col 数据  class:openpyxl.cell
        """
        return self.get_rows(row)[col]

    def _typestrip(self, args):
        """
        类型判断保护用
        :param args:
        :return:
        """
        return args.strip() if type(args) is str else args

    # def _write_cell(self,colsName,index):
    #     """
    #     写入保护用
    #     :param colsName: str
    #     :param index: int 0开始
    #     :return:
    #     """
    #     if isinstance(colsName, str) and isinstance(index, int):
    #         s1 =colsName+str(index)
    #         return s1
    #     else:
    #         raise Exception("colsName为str类型,index为int类型")
    #
    #
    # def write_input_data(self):
    #     wb =openpyxl.load_workbook(self.xlsx_path)
    #     sheet = wb.sheetnames[0]
    #     ws =wb[sheet]
    #     return ws
    #
    #
    # def load_workbook(self):
    #     wb = openpyxl.load_workbook(self.xlsx_path)
    #     return wb